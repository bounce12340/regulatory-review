#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cost Tracker for Regulatory Review Projects
法規審查專案費用追蹤系統

Tracks expenses, budgets, and forecasts for regulatory submissions.
Usage:
  python cost_tracker.py add --project fenogal --category government_fees --subcategory review_fee --amount 15000 --currency TWD
  python cost_tracker.py list --project fenogal
  python cost_tracker.py report --project fenogal
  python cost_tracker.py budget --project fenogal --type drug_registration_extension --total 500000
  python cost_tracker.py export --project fenogal --format excel
  python cost_tracker.py forecast --project fenogal
"""

import io
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.text import Text
    from rich.progress import Progress, BarColumn, TextColumn
    from rich import box
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

console = Console(highlight=False) if RICH_AVAILABLE else None

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

SCRIPT_DIR   = Path(__file__).parent
CONFIG_DIR   = SCRIPT_DIR.parent / "config"
COSTS_DIR    = Path.home() / "productivity" / "costs"
CONFIG_FILE  = CONFIG_DIR / "cost_categories.yaml"

# ---------------------------------------------------------------------------
# Config loader
# ---------------------------------------------------------------------------

_CONFIG_CACHE: Optional[Dict] = None

def load_config() -> Dict:
    global _CONFIG_CACHE
    if _CONFIG_CACHE is not None:
        return _CONFIG_CACHE
    if YAML_AVAILABLE and CONFIG_FILE.exists():
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            _CONFIG_CACHE = yaml.safe_load(f)
    else:
        # Minimal fallback config
        _CONFIG_CACHE = {
            "exchange_rates": {"TWD": 1.0, "USD": 31.5, "EUR": 34.2},
            "budget_alert_threshold": 0.85,
            "categories": {},
            "project_budgets": {},
        }
    return _CONFIG_CACHE


def to_twd(amount: float, currency: str) -> float:
    """Convert any supported currency to TWD."""
    rates = load_config().get("exchange_rates", {"TWD": 1.0, "USD": 31.5, "EUR": 34.2})
    rate  = rates.get(currency.upper(), 1.0)
    return amount * rate


def format_currency(amount_twd: float, currency: str = "TWD") -> str:
    """Format amount in the given currency."""
    rates = load_config().get("exchange_rates", {"TWD": 1.0, "USD": 31.5, "EUR": 34.2})
    rate  = rates.get(currency.upper(), 1.0)
    value = amount_twd / rate
    symbols = {"TWD": "NT$", "USD": "$", "EUR": "€"}
    sym = symbols.get(currency.upper(), currency)
    return f"{sym}{value:,.0f}"


# ---------------------------------------------------------------------------
# Data layer
# ---------------------------------------------------------------------------

def get_project_cost_file(project: str) -> Path:
    path = COSTS_DIR / project / "costs.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def load_costs(project: str) -> Dict:
    path = get_project_cost_file(project)
    if path.exists():
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"project": project, "expenses": [], "budgets": {}, "created": datetime.now().isoformat()}


def save_costs(project: str, data: Dict) -> Path:
    path = get_project_cost_file(project)
    data["last_updated"] = datetime.now().isoformat()
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


# ---------------------------------------------------------------------------
# Expense operations
# ---------------------------------------------------------------------------

def add_expense(
    project: str,
    category: str,
    subcategory: str,
    amount: float,
    currency: str,
    description: str = "",
    vendor: str = "",
    date: Optional[str] = None,
    invoice_ref: str = "",
) -> Dict:
    """Add an expense entry to a project."""
    data    = load_costs(project)
    expense = {
        "id":          f"EXP-{len(data['expenses']) + 1:04d}",
        "date":        date or datetime.now().strftime("%Y-%m-%d"),
        "category":    category,
        "subcategory": subcategory,
        "amount":      amount,
        "currency":    currency.upper(),
        "amount_twd":  to_twd(amount, currency),
        "description": description,
        "vendor":      vendor,
        "invoice_ref": invoice_ref,
        "added_at":    datetime.now().isoformat(),
    }
    data["expenses"].append(expense)
    save_costs(project, data)
    return expense


def set_budget(project: str, project_type: str, total_twd: float) -> Dict:
    """Set project budget based on project type template."""
    data   = load_costs(project)
    config = load_config()
    template = config.get("project_budgets", {}).get(project_type, {})

    allocation = template.get("category_allocation", {})
    budget = {
        "project_type":    project_type,
        "total_twd":       total_twd,
        "set_date":        datetime.now().isoformat(),
        "category_budgets": {
            cat: round(total_twd * pct)
            for cat, pct in allocation.items()
        },
    }
    data["budgets"][project_type] = budget
    data["active_budget"]         = project_type
    save_costs(project, data)
    return budget


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def compute_summary(project: str) -> Dict:
    """Compute spending summary for a project."""
    data     = load_costs(project)
    expenses = data.get("expenses", [])

    total_twd = sum(e["amount_twd"] for e in expenses)

    by_category: Dict[str, float] = {}
    by_month:    Dict[str, float] = {}

    for e in expenses:
        cat = e.get("category", "miscellaneous")
        by_category[cat] = by_category.get(cat, 0) + e["amount_twd"]
        month = e["date"][:7]  # YYYY-MM
        by_month[month] = by_month.get(month, 0) + e["amount_twd"]

    # Budget info
    active_type = data.get("active_budget")
    budget_info = data.get("budgets", {}).get(active_type, {}) if active_type else {}
    budget_total = budget_info.get("total_twd", 0)
    budget_used_pct = (total_twd / budget_total * 100) if budget_total else 0

    # Category budget vs actual
    cat_budgets = budget_info.get("category_budgets", {})
    cat_analysis = {}
    for cat, spent in by_category.items():
        budgeted = cat_budgets.get(cat, 0)
        cat_analysis[cat] = {
            "spent":     spent,
            "budgeted":  budgeted,
            "remaining": budgeted - spent,
            "pct_used":  (spent / budgeted * 100) if budgeted else 0,
        }

    return {
        "project":          project,
        "total_expenses":   len(expenses),
        "total_twd":        total_twd,
        "budget_total":     budget_total,
        "budget_used_pct":  budget_used_pct,
        "budget_remaining": budget_total - total_twd,
        "by_category":      by_category,
        "by_month":         dict(sorted(by_month.items())),
        "cat_analysis":     cat_analysis,
        "active_budget":    active_type,
        "alert":            budget_used_pct >= (load_config().get("budget_alert_threshold", 0.85) * 100),
    }


def forecast_costs(project: str, target_months: int = 3) -> Dict:
    """Forecast future costs based on historical burn rate."""
    data     = load_costs(project)
    expenses = data.get("expenses", [])

    if not expenses:
        return {"error": "No expense data available for forecasting"}

    # Group by month
    by_month: Dict[str, float] = {}
    for e in expenses:
        month = e["date"][:7]
        by_month[month] = by_month.get(month, 0) + e["amount_twd"]

    monthly_amounts = list(by_month.values())
    if not monthly_amounts:
        return {"error": "No monthly data"}

    avg_monthly  = sum(monthly_amounts) / len(monthly_amounts)
    max_monthly  = max(monthly_amounts)
    min_monthly  = min(monthly_amounts)

    summary = compute_summary(project)
    budget_remaining = summary.get("budget_remaining", 0)
    months_until_exhausted = (budget_remaining / avg_monthly) if avg_monthly > 0 else float('inf')

    return {
        "avg_monthly_spend":      avg_monthly,
        "max_monthly_spend":      max_monthly,
        "min_monthly_spend":      min_monthly,
        "forecast_next_months":   target_months,
        "forecast_total":         avg_monthly * target_months,
        "forecast_optimistic":    min_monthly * target_months,
        "forecast_pessimistic":   max_monthly * target_months,
        "months_data_points":     len(monthly_amounts),
        "budget_months_remaining": months_until_exhausted,
        "by_month_history":       by_month,
    }


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_to_csv(project: str, output_path: Optional[str] = None) -> Path:
    """Export expenses to CSV."""
    data = load_costs(project)
    dest = Path(output_path) if output_path else COSTS_DIR / project / f"costs-{datetime.now().strftime('%Y%m%d')}.csv"
    dest.parent.mkdir(parents=True, exist_ok=True)

    expenses = data.get("expenses", [])
    if not expenses:
        dest.write_text("No expenses found\n", encoding="utf-8")
        return dest

    # Write CSV manually (no pandas dependency)
    fields = ["id", "date", "category", "subcategory", "amount", "currency", "amount_twd", "description", "vendor", "invoice_ref"]
    lines  = [",".join(fields)]
    for e in expenses:
        row = [str(e.get(f, "")) for f in fields]
        lines.append(",".join(f'"{v}"' if "," in v else v for v in row))
    dest.write_text("\n".join(lines), encoding="utf-8")
    return dest


def export_to_excel(project: str, output_path: Optional[str] = None) -> Path:
    """Export expenses and summary to Excel."""
    dest = Path(output_path) if output_path else COSTS_DIR / project / f"costs-{datetime.now().strftime('%Y%m%d')}.xlsx"
    dest.parent.mkdir(parents=True, exist_ok=True)

    if not OPENPYXL_AVAILABLE:
        _warn("openpyxl not installed — falling back to CSV")
        return export_to_csv(project, str(dest.with_suffix(".csv")))

    data     = load_costs(project)
    expenses = data.get("expenses", [])
    summary  = compute_summary(project)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Expenses ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Expenses"
    headers = ["ID", "Date", "Category", "Subcategory", "Amount", "Currency", "Amount (TWD)", "Description", "Vendor", "Invoice Ref"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    for row_idx, e in enumerate(expenses, 2):
        ws1.cell(row=row_idx, column=1,  value=e.get("id", ""))
        ws1.cell(row=row_idx, column=2,  value=e.get("date", ""))
        ws1.cell(row=row_idx, column=3,  value=e.get("category", ""))
        ws1.cell(row=row_idx, column=4,  value=e.get("subcategory", ""))
        ws1.cell(row=row_idx, column=5,  value=e.get("amount", 0))
        ws1.cell(row=row_idx, column=6,  value=e.get("currency", ""))
        ws1.cell(row=row_idx, column=7,  value=e.get("amount_twd", 0))
        ws1.cell(row=row_idx, column=8,  value=e.get("description", ""))
        ws1.cell(row=row_idx, column=9,  value=e.get("vendor", ""))
        ws1.cell(row=row_idx, column=10, value=e.get("invoice_ref", ""))

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value=f"Project: {project}").font = openpyxl.styles.Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=4, column=1, value="Total Expenses (TWD)").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=4, column=2, value=summary["total_twd"])
    ws2.cell(row=5, column=1, value="Budget (TWD)").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=5, column=2, value=summary["budget_total"])
    ws2.cell(row=6, column=1, value="Budget Used %").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=6, column=2, value=f"{summary['budget_used_pct']:.1f}%")

    ws2.cell(row=8, column=1, value="Category").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=8, column=2, value="Spent (TWD)").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=8, column=3, value="Budgeted (TWD)").font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=8, column=4, value="% Used").font = openpyxl.styles.Font(bold=True)

    for row_idx, (cat, vals) in enumerate(summary["cat_analysis"].items(), 9):
        ws2.cell(row=row_idx, column=1, value=cat)
        ws2.cell(row=row_idx, column=2, value=vals["spent"])
        ws2.cell(row=row_idx, column=3, value=vals["budgeted"])
        ws2.cell(row=row_idx, column=4, value=f"{vals['pct_used']:.1f}%")

    wb.save(dest)
    return dest


# ---------------------------------------------------------------------------
# Rich display
# ---------------------------------------------------------------------------

def _warn(msg: str):
    if RICH_AVAILABLE and console:
        console.print(f"[bold yellow]Warning:[/bold yellow] {msg}")
    else:
        print(f"Warning: {msg}")


def render_expense_list(project: str, expenses: List[Dict]):
    if not RICH_AVAILABLE or not console:
        print(f"\nExpenses for {project}:")
        for e in expenses:
            print(f"  {e['id']} | {e['date']} | {e['category']}/{e['subcategory']} | {e['currency']} {e['amount']:,.0f} ({e['amount_twd']:,.0f} TWD) | {e.get('description','')}")
        return

    table = Table(
        title=f"Expenses — {project.upper()}",
        box=box.ROUNDED,
        show_lines=True,
        border_style="blue",
    )
    table.add_column("ID",          style="dim",         width=10)
    table.add_column("Date",        style="cyan",        width=12)
    table.add_column("Category",    style="bold",        min_width=16)
    table.add_column("Subcategory", style="dim",         min_width=14)
    table.add_column("Amount",      justify="right",     min_width=14)
    table.add_column("TWD",         justify="right",     min_width=12, style="bold green")
    table.add_column("Description", style="dim",         min_width=20)

    for e in expenses:
        table.add_row(
            e["id"],
            e["date"],
            e["category"],
            e.get("subcategory", ""),
            f"{e['currency']} {e['amount']:,.0f}",
            f"NT${e['amount_twd']:,.0f}",
            e.get("description", "") or e.get("vendor", ""),
        )
    console.print(table)


def render_summary(summary: Dict):
    if not RICH_AVAILABLE or not console:
        print(f"\nCost Summary — {summary['project'].upper()}")
        print(f"  Total Spent:  NT${summary['total_twd']:,.0f}")
        print(f"  Budget:       NT${summary['budget_total']:,.0f}")
        print(f"  Budget Used:  {summary['budget_used_pct']:.1f}%")
        if summary.get("alert"):
            print("  [!] BUDGET ALERT: Approaching limit!")
        return

    # Header panel
    alert_style = "bold red" if summary["alert"] else "bold green"
    pct = summary["budget_used_pct"]
    header = Text()
    header.append(f"Project:   ", style="bold")
    header.append(f"{summary['project'].upper()}\n", style="bold cyan")
    header.append(f"Expenses:  {summary['total_expenses']} records\n")
    header.append(f"Total:     ")
    header.append(f"NT${summary['total_twd']:,.0f}\n", style="bold green")
    if summary["budget_total"]:
        header.append(f"Budget:    NT${summary['budget_total']:,.0f}\n")
        header.append(f"Remaining: NT${summary['budget_remaining']:,.0f}\n",
                      style="bold red" if summary["budget_remaining"] < 0 else "bold green")
        if summary["alert"]:
            header.append("\n⚠  BUDGET ALERT — Approaching limit!", style="bold red blink")
    console.print(Panel(header, title="[bold]Cost Summary[/bold]", border_style="blue"))

    # Budget progress bar
    if summary["budget_total"]:
        bar_style = "red" if pct >= 85 else "yellow" if pct >= 60 else "green"
        with Progress(
            TextColumn("[bold]Budget Used[/bold]"),
            BarColumn(bar_width=40, complete_style=bar_style),
            TextColumn(f"{pct:.1f}% (NT${summary['total_twd']:,.0f} / NT${summary['budget_total']:,.0f})"),
            console=console,
            transient=False,
        ) as progress:
            task = progress.add_task("", total=100, completed=min(pct, 100))
            progress.update(task)

    # Category breakdown
    if summary["cat_analysis"]:
        table = Table(title="By Category", box=box.ROUNDED, border_style="blue")
        table.add_column("Category",   style="bold",  min_width=20)
        table.add_column("Spent",      justify="right", min_width=14)
        table.add_column("Budgeted",   justify="right", min_width=14)
        table.add_column("Remaining",  justify="right", min_width=14)
        table.add_column("Used %",     justify="right", min_width=8)

        for cat, vals in summary["cat_analysis"].items():
            pct_cat = vals["pct_used"]
            pct_style = "bold red" if pct_cat >= 90 else "bold yellow" if pct_cat >= 70 else "green"
            rem_style = "bold red" if vals["remaining"] < 0 else "green"
            table.add_row(
                cat,
                f"NT${vals['spent']:,.0f}",
                f"NT${vals['budgeted']:,.0f}" if vals["budgeted"] else "—",
                Text(f"NT${vals['remaining']:,.0f}", style=rem_style),
                Text(f"{pct_cat:.0f}%", style=pct_style),
            )
        console.print(table)

    # Monthly breakdown
    if summary["by_month"]:
        table2 = Table(title="Monthly Spending", box=box.ROUNDED, border_style="dim")
        table2.add_column("Month", style="cyan", min_width=10)
        table2.add_column("Amount (TWD)", justify="right", min_width=16)
        for month, amt in summary["by_month"].items():
            table2.add_row(month, f"NT${amt:,.0f}")
        console.print(table2)


def render_forecast(forecast: Dict):
    if "error" in forecast:
        if RICH_AVAILABLE and console:
            console.print(f"[bold red]Forecast error:[/bold red] {forecast['error']}")
        else:
            print(f"Forecast error: {forecast['error']}")
        return

    if not RICH_AVAILABLE or not console:
        print(f"\nCost Forecast ({forecast['forecast_next_months']} months):")
        print(f"  Avg monthly:      NT${forecast['avg_monthly_spend']:,.0f}")
        print(f"  Forecast total:   NT${forecast['forecast_total']:,.0f}")
        print(f"  Optimistic:       NT${forecast['forecast_optimistic']:,.0f}")
        print(f"  Pessimistic:      NT${forecast['forecast_pessimistic']:,.0f}")
        if forecast.get("budget_months_remaining", float("inf")) != float("inf"):
            print(f"  Budget runway:    {forecast['budget_months_remaining']:.1f} months")
        return

    text = Text()
    text.append(f"Forecast Period:   ", style="bold")
    text.append(f"{forecast['forecast_next_months']} months\n")
    text.append(f"Data Points:       {forecast['months_data_points']} months of history\n", style="dim")
    text.append(f"\nAvg Monthly Burn:  ")
    text.append(f"NT${forecast['avg_monthly_spend']:,.0f}\n", style="bold cyan")
    text.append(f"Forecast (avg):    ")
    text.append(f"NT${forecast['forecast_total']:,.0f}\n", style="bold yellow")
    text.append(f"Optimistic:        NT${forecast['forecast_optimistic']:,.0f}\n", style="green")
    text.append(f"Pessimistic:       NT${forecast['forecast_pessimistic']:,.0f}\n", style="red")

    runway = forecast.get("budget_months_remaining", float("inf"))
    if runway != float("inf"):
        runway_style = "bold red" if runway < 2 else "bold yellow" if runway < 4 else "bold green"
        text.append(f"\nBudget Runway:     ")
        text.append(f"{runway:.1f} months remaining\n", style=runway_style)

    console.print(Panel(text, title="[bold cyan]Cost Forecast[/bold cyan]", border_style="cyan"))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def cmd_add(args: argparse.Namespace):
    expense = add_expense(
        project     = args.project,
        category    = args.category,
        subcategory = args.subcategory,
        amount      = args.amount,
        currency    = args.currency,
        description = args.description or "",
        vendor      = args.vendor or "",
        date        = args.date,
        invoice_ref = args.invoice or "",
    )
    if RICH_AVAILABLE and console:
        console.print(
            f"[bold green]✓ Added[/bold green] {expense['id']}: "
            f"{args.currency} {args.amount:,.0f} "
            f"([bold]NT${expense['amount_twd']:,.0f}[/bold]) — "
            f"{args.category}/{args.subcategory}"
        )
    else:
        print(f"Added {expense['id']}: {args.currency} {args.amount:,.0f} = NT${expense['amount_twd']:,.0f}")


def cmd_list(args: argparse.Namespace):
    data     = load_costs(args.project)
    expenses = data.get("expenses", [])
    if args.category:
        expenses = [e for e in expenses if e.get("category") == args.category]
    render_expense_list(args.project, expenses)


def cmd_report(args: argparse.Namespace):
    summary = compute_summary(args.project)
    render_summary(summary)


def cmd_budget(args: argparse.Namespace):
    budget = set_budget(args.project, args.type, args.total)
    if RICH_AVAILABLE and console:
        console.print(
            f"[bold green]✓ Budget set[/bold green] for [cyan]{args.project}[/cyan]: "
            f"NT${args.total:,.0f} ({args.type})"
        )
        for cat, amt in budget["category_budgets"].items():
            console.print(f"   {cat}: NT${amt:,.0f}")
    else:
        print(f"Budget set: NT${args.total:,.0f} for {args.project} ({args.type})")


def cmd_forecast(args: argparse.Namespace):
    result = forecast_costs(args.project, args.months)
    render_forecast(result)


def cmd_export(args: argparse.Namespace):
    fmt = (args.format or "csv").lower()
    if fmt == "excel":
        path = export_to_excel(args.project, args.output)
    else:
        path = export_to_csv(args.project, args.output)
    if RICH_AVAILABLE and console:
        console.print(f"[bold green]✓ Exported[/bold green] → {path}")
    else:
        print(f"Exported → {path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Cost Tracker — Regulatory Review Projects / 法規審查費用追蹤")
    sub    = parser.add_subparsers(dest="command", required=True)

    # ── add ──────────────────────────────────────────────────────────────
    p_add = sub.add_parser("add", help="Add expense / 新增費用")
    p_add.add_argument("--project",     required=True)
    p_add.add_argument("--category",    required=True, help="e.g. government_fees")
    p_add.add_argument("--subcategory", required=True, help="e.g. review_fee")
    p_add.add_argument("--amount",      required=True, type=float)
    p_add.add_argument("--currency",    default="TWD",  choices=["TWD", "USD", "EUR"])
    p_add.add_argument("--description", help="Expense description / 費用說明")
    p_add.add_argument("--vendor",      help="Vendor / payee name / 廠商")
    p_add.add_argument("--date",        help="YYYY-MM-DD (defaults to today)")
    p_add.add_argument("--invoice",     help="Invoice / receipt reference")

    # ── list ─────────────────────────────────────────────────────────────
    p_list = sub.add_parser("list", help="List expenses / 列出費用")
    p_list.add_argument("--project",  required=True)
    p_list.add_argument("--category", help="Filter by category")

    # ── report ───────────────────────────────────────────────────────────
    p_rep = sub.add_parser("report", help="Cost summary report / 費用摘要報告")
    p_rep.add_argument("--project", required=True)

    # ── budget ───────────────────────────────────────────────────────────
    p_bud = sub.add_parser("budget", help="Set project budget / 設定預算")
    p_bud.add_argument("--project", required=True)
    p_bud.add_argument("--type",    required=True,
                       choices=["drug_registration_extension", "food_registration", "medical_device_registration"])
    p_bud.add_argument("--total",   required=True, type=float, help="Total budget in TWD")

    # ── forecast ─────────────────────────────────────────────────────────
    p_fore = sub.add_parser("forecast", help="Forecast future costs / 費用預測")
    p_fore.add_argument("--project", required=True)
    p_fore.add_argument("--months",  type=int, default=3)

    # ── export ───────────────────────────────────────────────────────────
    p_exp = sub.add_parser("export", help="Export to CSV or Excel / 匯出")
    p_exp.add_argument("--project", required=True)
    p_exp.add_argument("--format",  choices=["csv", "excel"], default="csv")
    p_exp.add_argument("--output",  help="Output file path")

    args = parser.parse_args()
    dispatch = {
        "add":      cmd_add,
        "list":     cmd_list,
        "report":   cmd_report,
        "budget":   cmd_budget,
        "forecast": cmd_forecast,
        "export":   cmd_export,
    }
    dispatch[args.command](args)
    return 0


if __name__ == "__main__":
    sys.exit(main())
